import random
import openpyxl

def getData():
    workbook = openpyxl.load_workbook("./uids.xlsx", data_only=True)
    list = []
    sheet = workbook[workbook.sheetnames[0]]
    for row in range(1, 35):
        if sheet.cell(row= row, column=1).value == "":
            continue
        else:
            list.append(sheet.cell(row=row, column=1).value)
    return list



def getResult(list, num):
    result = random.sample(list, num)
    return result



if __name__ == '__main__':
    list = getData()
    res = getResult(list, 2)
    print(res)