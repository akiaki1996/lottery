from tkinter import *
import random
import openpyxl

is_run = False

def getPeopleList():
    workbook = openpyxl.load_workbook("./uids.csv", data_only=True)
    list = []
    sheet = workbook[workbook.sheetnames[0]]
    for row in range(2, sheet.max_row+1):
        if sheet.cell(row= row, column=2).value == "":
            continue
        else:
            list.append(sheet.cell(row=row, column=2).value)

    return list

def randomRun(list):
    global is_run
    if is_run:
        return
    is_run = True
    num = insert_point()
    if num:
        num = int(num)
        randomResult(list,num)
    else:
        is_run = False
    


    def insert_point():
        var = input.get()
        return var
    

    def randomResult(list, num):
        global is_run
        result = random.sample(list, num)
        var.set(result)
        if is_run:
            window.after(100, randomResult,list, num)

    def finalResult():
        global is_run
        is_run = False
    

    if __name__ == '__main__':
        window = Tk()
        window.getmetry('500x290+250+150')
        window.resizable(0, 0)
        window.title('陈琳问卷调查抽奖0829')
        list = getPeopleList()


        var = StringVar() #initialize a string var, for rolling presenting results

        noteLabel = Label(testvariable = var)
        noteLabel.place(anchor=NW, x = 120, y = 120)
        input = Entry(window,show=None) #input frame
        input.place(anchor=NW,x = 240,y = 30)

        resultLabel = Label(textvariable=var)
        resultLabel.place(anchor=NW,x=150,y=100)

        startBt = Button(text="start", command=lambda:randomRun(list=list)) #start
        confirmBt = Button(text="confirm", command=lambda:finalResult()) #pause
        startBt.place(anchor=NW, x = 200, y = 180)
        confirmBt.place(anchor=NW, x=260,y=180)

        window.mainloop()




